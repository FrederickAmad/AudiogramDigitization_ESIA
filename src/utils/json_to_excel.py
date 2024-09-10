import json
import os
import sys
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def format_cell_value(value):
    if isinstance(value, list):
        return ', '.join(map(str, value))  # Join list items with commas
    elif isinstance(value, dict):
        return json.dumps(value)  # Convert dict to JSON string
    else:
        return value  # Return as is for other types

def json_files_to_excel(input_folder, output_excel):
    workbook = Workbook()
    workbook.remove(workbook.active)  # Remove the default sheet

    for filename in os.listdir(input_folder):
        if filename.endswith('.json'):
            file_path = os.path.join(input_folder, filename)
            with open(file_path, 'r') as json_file:
                try:
                    data = json.load(json_file)
                    if not isinstance(data, list):
                        data = [data]

                    # Create a new sheet for each JSON file
                    sheet_name = os.path.splitext(filename)[0][:31]  # Excel sheet names are limited to 31 characters
                    sheet = workbook.create_sheet(title=sheet_name)

                    # Get all unique keys
                    all_keys = set()
                    for item in data:
                        all_keys.update(item.keys())

                    # Sort the keys to ensure consistent column order
                    headers = sorted(all_keys)

                    # Write headers
                    for col, header in enumerate(headers, start=1):
                        sheet.cell(row=1, column=col, value=header)

                    # Write data
                    for row, item in enumerate(data, start=2):
                        for col, key in enumerate(headers, start=1):
                            value = item.get(key, '')
                            formatted_value = format_cell_value(value)
                            sheet.cell(row=row, column=col, value=formatted_value)

                    # Adjust column widths
                    for col in range(1, len(headers) + 1):
                        sheet.column_dimensions[get_column_letter(col)].auto_size = True

                    print(f"Processed: {filename}")
                except json.JSONDecodeError:
                    print(f"Error reading {filename}. Skipping this file.")
                except Exception as e:
                    print(f"Unexpected error processing {filename}: {str(e)}. Skipping this file.")

    if len(workbook.sheetnames) == 0:
        print("No valid JSON data found in the input folder.")
        return

    try:
        workbook.save(output_excel)
        print(f"Conversion complete. Excel file saved as {output_excel}")
    except PermissionError:
        print(f"Error: Unable to save the Excel file. The file may be open in another program.")
    except Exception as e:
        print(f"Error saving Excel file: {str(e)}")

def main():
    # Argument placeholders
    if len(sys.argv) != 3:
        print("Usage: python script.py <input_folder> <output_excel>")
        print("Converts JSON files in the input folder to an Excel file.")
        print("Example: python script.py /path/to/json_folder /path/to/output.xlsx")
        return

    input_folder = sys.argv[1]  # First argument: input folder containing JSON files
    output_excel = sys.argv[2]  # Second argument: output Excel file path

    # Check if the input folder exists
    if not os.path.exists(input_folder):
        print(f"Error: The input folder '{input_folder}' does not exist.")
        return

    json_files_to_excel(input_folder, output_excel)

if __name__ == "__main__":
    main()